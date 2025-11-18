import openpyxl
import json
from typing import Dict, Any, Optional, List, Tuple
from dataclasses import dataclass
from pathlib import Path
import pandas as pd

from src.utils.logger_config import logger    # ✅ Professional logger




@dataclass
class FieldConfig:
    """Configuration for each field to extract"""
    name: str
    positions: List[Tuple[int, int]]


def make_field_range(name: str, rows: range, col: int) -> FieldConfig:
    return FieldConfig(name, [(r, col) for r in rows])


def make_middle(row: int, col: int) -> List[FieldConfig]:
    fields = []
    names = ["customer", "manufacturer", "sevensix"]

    for i in range(3):
        rows = range(row + i * 5, row + i * 5 + 5)
        fields.extend([
            FieldConfig(f"ai_{names[i]}_who", [(r, col) for r in rows]),
            FieldConfig(f"ai_{names[i]}_what", [(r, col + 1) for r in rows]),
            FieldConfig(f"ai_{names[i]}_when", [(r, col + 5) for r in rows]),
        ])
    return fields


def make_case_fields(case_no: int, start_row: int, start_col: int = 6) -> List[FieldConfig]:
    fields = [
        ("customer_name", (start_row, start_col)),
        ("customer_representative", (start_row, start_col+1)),
        ("occurrence_date", (start_row, start_col+2)),
        ("branch", (start_row, start_col+3)),
        ("application", (start_row+1, start_col)),
        ("manufacturer", (start_row+2, start_col)),
        ("product", (start_row+3, start_col)),
        ("competitive", (start_row+4, start_col)),
        ("amount", (start_row+5, start_col)),
        ("order_month", (start_row+6, start_col)),
        ("budget", (start_row+7, start_col)),
        ("probability", (start_row+8, start_col)),
    ]
    return [FieldConfig(f"case{case_no}_{fname}", [pos]) for fname, pos in fields]



class ExcelDataExtractor:
    """Optimized Excel data extractor with ETL logging"""
    
    def __init__(self):
        logger.info("Initializing ExcelDataExtractor...")
        
        # Basic fields
        self.fields = [
            FieldConfig("reporter_name", [(2,4), (2,5), (2,6)]),
            FieldConfig("date", [(3,4), (3,5)]),
            FieldConfig("location", [(4,4), (4,5)]),

            make_field_range("customer_company_name", range(7, 11), 4),
            make_field_range("customer_department", range(7, 12), 5),
            make_field_range("customer_full_name", range(7, 12), 6),
            FieldConfig("customer_number", [(11,4)]),

            make_field_range("manufacturer_company_name", range(7, 11), 7),
            make_field_range("manufacturer_department", range(7, 11), 8),
            make_field_range("manufacturer_full_name", range(7, 11), 9),
            FieldConfig("sevensix", [(7,10),(8,10),(9,10),(10,10)]),
        ]

        logger.success("ExcelDataExtractor initialized successfully.")

    def adjust_fields_by_dimension(self, ws):
        """Dynamically add fields based on Excel template size"""

        if ws.max_row == 71:
            # logger.info("Using template: 71-row layout")

            self.fields += [
                FieldConfig("purpose", [(12,4),(13,4),(16,4)]),
                FieldConfig("free_description", [(17,4),(18,4)]),
                FieldConfig("associated_customer_name", [(35,4)]),
                FieldConfig("competitive_information", [(36,4)]),
            ]
            self.fields.extend(make_middle(20, 5))

            self.fields += make_case_fields(1, 38)
            self.fields += make_case_fields(2, 48)
            self.fields += make_case_fields(3, 58)

        elif ws.max_row == 97:
            # logger.info("Using template: 97-row layout")

            self.fields += [
                FieldConfig("annual_project_name", [(17,4),(17,5),(17,6)]),
                FieldConfig("annual_project_schedule", [(18,4),(18,5),(18,6)]),
                FieldConfig("annual_project_budget", [(19,4),(19,5),(19,6)]),
                FieldConfig("free_description", [(20,4),(20,5)]),
                FieldConfig("associated_customer_name", [(46,4)]),
                FieldConfig("competitive_information", [(47,4)]),
            ]

            self.fields.extend(make_middle(31, 5))

            self.fields += make_case_fields(1, 49)
            self.fields += make_case_fields(2, 59)
            self.fields += make_case_fields(3, 69)
            self.fields += make_case_fields(4, 79)
            self.fields += make_case_fields(5, 89)

        else:
            logger.warning(f"Unknown template size: {ws.max_row} rows. Extraction may be incomplete.")

    def extract_from_file(self, file_path: str) -> Dict[str, Any]:
        """Extract structured data from Excel"""
        # logger.info(f"Extracting file: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            logger.debug(f"Worksheet loaded: {ws.title}, Size: {ws.max_row}x{ws.max_column}")

            self.adjust_fields_by_dimension(ws)

            extracted, stats = {}, {"found": 0, "not_found": 0}

            for field in self.fields:
                value = self._extract_field(ws, field)
                extracted[field.name] = value

                if value:
                    stats["found"] += 1
                else:
                    stats["not_found"] += 1
                    # logger.debug(f"Field not found: {field.name} positions={field.positions}")

            results = {
                "data": extracted,
                "file_info": {
                    "filename": Path(file_path).name,
                    "worksheet": ws.title,
                    "dimensions": f"{ws.max_row}x{ws.max_column}",
                },
                "extraction_stats": {
                    **stats,
                    "total_fields": len(self.fields),
                    "success_rate": f"{(stats['found']/len(self.fields))*100:.1f}%",
                },
            }

            # logger.success(
            #     f"Extraction complete: {Path(file_path).name} | "
            #     f"Success={stats['found']} Missing={stats['not_found']} "
            #     f"({results['extraction_stats']['success_rate']})"
            # )

            return results

        except Exception as e:
            logger.error(f"Extraction failed for {file_path}: {e}")
            return {"error": str(e)}

    def _extract_field(self, ws, field: FieldConfig) -> Optional[str]:
        """Extract text from fixed cell positions"""
        values = [
            str(ws.cell(row=r, column=c).value).strip()
            for r, c in field.positions
            if ws.cell(row=r, column=c).value not in (None, "")
        ]

        return "\n".join(values) if len(values) > 1 else (values[0] if values else None)

    def save_results(self, results: Dict[str, Any], output_file: str):
        try:
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)

            logger.success(f"Saved extraction JSON → {output_file}")

        except Exception as e:
            logger.error(f"Failed to save output file ({output_file}): {e}")

    def batch_process(self, files: List[str], level="1", origin="s3_bucket", output_dir="outputs"):
        logger.info(f"Batch processing {len(files)} Excel files...")

        Path(output_dir).mkdir(exist_ok=True)
        all_rows = []

        for i, file in enumerate(files, 1):
            # logger.info(f"[{i}/{len(files)}] Processing: {Path(file).name}")

            result = self.extract_from_file(file)

            if result and "data" in result:
                row = result["data"].copy()
                row["source"] = file
                row["level"] = level
                row["origin"] = origin
                all_rows.append(row)
            else:
                logger.warning(f"Skipping file due to extraction errors: {file}")

        df = pd.DataFrame(all_rows)

        logger.success(f"Batch processing completed. Extracted rows: {len(df)}")

        return df

    def list_excel_files(self, source_dir: str, extensions=("*.xlsx", "*.xlsm")) -> List[str]:
        p = Path(source_dir)
        files = []

        for ext in extensions:
            files.extend(p.rglob(ext))

        logger.info(f"Found {len(files)} Excel files in {source_dir}")

        return [str(f) for f in files]
