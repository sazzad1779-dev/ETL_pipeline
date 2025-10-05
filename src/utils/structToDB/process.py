import os
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

class StructToProcess:
    def __init__(self,level:str="1",origin:str="s3_bucket"):
        self.level =level
        self.origin = origin

    def _get_file_size_mb(self, filepath):
        return os.path.getsize(filepath) / (1024 * 1024)

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        return df

    def _rename_duplicate_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        seen = defaultdict(int)
        new_columns = []
        for col in df.columns:
            if seen[col]:
                new_col = f"{col}_{seen[col]}"
            else:
                new_col = col
            seen[col] += 1
            new_columns.append(new_col)
        df.columns = new_columns
        return df

    # ---------------- Loaders ----------------
    def _load_with_pandas(self, file_path, table_name):
        encodings_to_try = ["utf-8", "cp932", "shift_jis"]
        for encoding in encodings_to_try:
            try:
                df = pd.read_csv(file_path, encoding=encoding, skiprows=7)
                df = df.drop_duplicates()
                df = self._rename_duplicate_columns(df)
                df = self._clean_dataframe(df)
                df["source"]=file_path
                df["level"]=self.level 
                df["origin"]=self.origin
                # print(f"[pandas] Loaded '{file_path}' → table '{table_name}' using encoding '{encoding}'")
                return {table_name: df}
            except UnicodeDecodeError as e:
                print(f" Encoding '{encoding}' failed for {file_path}: {e}")
            except Exception as e:
                print(f"❌ Unexpected error for {file_path}: {e}")
                return {}
        print(f"❌ All encoding attempts failed for {file_path}")
        return {}

    def _load_with_dask(self, file_path, table_name):
        import dask.dataframe as dd
        results = {}
        try:
            sample = pd.read_csv(file_path, nrows=5)
            col_dtypes = {col: "object" for col in sample.columns}
            ddf = dd.read_csv(file_path, blocksize="64MB", engine="python", dtype=col_dtypes)

            for i in range(ddf.npartitions):
                chunk = ddf.get_partition(i).compute()
                chunk = self._rename_duplicate_columns(chunk)
                chunk = self._clean_dataframe(chunk)
                results[f"{table_name}_part{i}"] = chunk

            print(f"[dask] Loaded '{file_path}' → {len(results)} partitions")
        except Exception as e:
            print(f"❌ Failed to load {file_path} with Dask: {e}")
        return results

    def _load_xlsx_with_pandas(self, file_path, table_name):
        results = {}
        try:
            wb = load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                merged_ranges = list(ws.merged_cells.ranges)

                for merged_range in merged_ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    value = ws.cell(row=min_row, column=min_col).value
                    if value is not None:
                        ws.unmerge_cells(str(merged_range))
                        for row in range(min_row, max_row + 1):
                            for col in range(min_col, max_col + 1):
                                ws.cell(row=row, column=col, value=value)

                df = pd.DataFrame(ws.values)
                header_row = self.find_header_row(df)
                df = df.drop(range(header_row)).reset_index(drop=True)
                df.columns = df.iloc[0]
                df = df.drop(0).reset_index(drop=True)
                df = self._rename_duplicate_columns(df)
                df = df.drop_duplicates()
                df = self._clean_dataframe(df)
                df["sheet_name"] = sheet_name
                df["source"]=file_path
                df["level"]=self.level 
                df["origin"]=self.origin

                full_table_name = f"{table_name}_{sheet_name}".lower()
                results[full_table_name] = df
                print(f"[xlsx] Loaded sheet '{sheet_name}' from '{file_path}' → table '{full_table_name}'")
        except Exception as e:
            print(f"❌ Failed to load Excel file {file_path}: {e}")
        return results

    def find_header_row(self, df: pd.DataFrame) -> int:
        non_empty_counts = df.notna().sum(axis=1)
        return int(non_empty_counts.idxmax())
